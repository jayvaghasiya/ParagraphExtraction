import fitz
import openpyxl
import argparse


def extract(args):
    # Open the PDF file
    doc = fitz.open(args.input_file)
    page_data_list = []

    # Iterate over each page in the document
    for page in doc:
        words = page.get_text_words()
        text_data = {}
        for index, word in enumerate(words):
            # find word block
            word_block = word[5]

            # Add word to corresponding block index
            if word_block not in text_data:
                text_data[word_block] = word[4]
            else:
                text_data[word_block] += " " + word[4]
        page_data_list.append(text_data)

    workbook = openpyxl.Workbook()
    # Iterate over each list and create a new sheet for it
    for i, data in enumerate(page_data_list):
        sheet_name = f"Page {i}"  # Use the list index as the sheet name
        worksheet = workbook.create_sheet(sheet_name)

        # Iterate over each line in the list and add it to the sheet
        for j, (block, line) in enumerate(data.items()):
            worksheet.cell(row=j + 1, column=1, value=line)
    workbook.save(args.output_file)


if __name__ == "__main__":
    # Parse command line arguments
    parser = argparse.ArgumentParser()
    parser.add_argument("-i", "--input_file", required=True, help="path to input file")
    parser.add_argument("-o", "--output_file", default="output.xlsx", help="path to output file")
    args = parser.parse_args()
    extract(args)
    print("success !")
